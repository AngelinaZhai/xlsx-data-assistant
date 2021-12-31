# Angelina Zhai
# May 28, 2019
# data_assistant.py
# Helps with merging and finding spaces in the xlsx file

# Importing packages
from Tkinter import *    # Tkinter
import tkMessageBox as box    # Message box
from tkFileDialog import askopenfilename     # Drop down menu for choosing files
import smtplib # Sending emails
from email.mime.text import MIMEText
import xlsxwriter
import openpyxl     # Reading xlsx files
from openpyxl import Workbook
from openpyxl import load_workbook

# Login Page
class Login:
     
     # Constructor
     # __init__(self, master)
     # @param: self:object, master:Tkinter
     # @return: none
     # Constructs the Login page with all the features
     def __init__(self, master):
          self.master = master
          master.title("Login")
          
          font1 = ("Cambria", 15)
          self.login_lbl = Label(master, text="Login", font = font1)
          self.login_lbl.grid(row = 0, column = 4, padx = 20)
          
          self.email_lbl = Label(master, text="Havergal Email:", justify = 'left')
          self.email_lbl.grid(row = 1, column = 0, columnspan = 2)
           
          self.email_ent = Entry(master)
          self.email_ent.grid(row = 1, column = 3, columnspan = 4, padx = 10, pady = 5)
          
          self.pw_lbl = Label(master, text = "Password:")
          self.pw_lbl.grid(row = 2, column = 1)
          
          self.pw_ent = Entry(master, show = "*")
          self.pw_ent.grid(row = 2, column = 3, columnspan = 4, padx = 10)
          
          self.enter_btn = Button(master, text = "Enter", command = self.enter)
          self.enter_btn.grid(row = 3, column = 4, padx = 5, pady = 10)
          
          self.message = Label(master, text = "", fg = 'red')
          self.message.grid(row = 4, column = 4)
          
          self.makeMenu()
     
     # makeMenu(self)
     # @param: self:object
     # @return: none
     # Assembles the menubar with its options
     def makeMenu(self):
          self.menubar = Menu(self.master)
          goto = Menu(self.menubar)
          goto.add_command(label = "Please log in first", command = self.msg)
          self.menubar.add_cascade(label = "Go to...", menu=goto)
          self.master.config(menu=self.menubar)

     # credentials(self)
     # @param: self:object
     # @return: emails:list, pws:list
     # Reads the 'login_credentials.txt' file and puts the data into two lists: emails and pws
     def credentials(self):
          f = open('login_credentials.txt')
          rawList = f.readlines()
          f.close()
          emails = []
          pws = []
          for i in range(0,len(rawList)):
               if '\n' in rawList[i]:
                    tmp = rawList[i]
                    rawList[i] = tmp[:len(tmp)-1]
               # if '\r' in rawList[i]:
               #      rawList[i] = rawList[i][0:len(rawList[i])-1]
               if '     ' in rawList[i]: 
                    line = rawList[i].split('     ')
                    emails.append(line[0])
                    pws.append(line[1])
          return emails, pws
     
     # clear(self)
     # @param: self:object
     # @return: none
     # Clears the entry boxes
     def clear(self):
          self.email_ent.delete(0,"end")
          self.pw_ent.delete(0,"end")
          
     # enter(self)
     # @param: self:object
     # @return: none
     # Compares the user in put with the credentials found in the login credentials; continues into program if credentials are correct, warning message box if not
    
    
     def enter(self):
          emails, pws = self.credentials()
          print pws
          self.user = self.email_ent.get()
          self.pw = self.pw_ent.get()
          if self.user in emails:
               ind = emails.index(self.user)
               #print "a=",pws[ind][0:len(pws[ind)-2])]
               if self.pw == pws[ind]:
                    self.message.configure(text = "Valid Login!", fg = 'black')
                    self.toChoose()
               else:
                    self.clear()
                    self.message.configure(text = "Invalid Login!")
          else:
               self.clear()
               self.message.configure(text = "Invalid Login!")
               
     # msg(self)
     # @param: self:object
     # @return: none
     # Shows warning message box
     def msg(self):
          box.showwarning("Login", "Please log in first!")
          
     # toChoose(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and move on to Choose
     def toChoose(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Choose(self.newWindow, self.user, self.pw)
        
        
# Choose Page
class Choose:
     
     # Constructor
     # __init__(self, master, user, password)
     # @param: self:object, master:Tkinter, user:str, password:str
     # @return: none
     # Constructs the Choose page with all features
     def __init__(self, master, user, password):
          self.master = master
          master.title("Choose action")
          self.user = user
          self.pw = password
          
          self.message = Label(master, text="What would you like to do?", justify = CENTER)
          self.message.pack(padx = 20, pady = 15)
          
          self.combine_btn = Button(master, text = "Combine files", command = lambda: self.newCombine())
          self.combine_btn.pack(side=LEFT, padx=5, pady = 10)
          
          self.compare_btn = Button(master, text = "Find Missing Info", command = lambda: self.newCompare())
          self.compare_btn.pack(side=LEFT, padx = 5)

          self.makeMenu()

     # makeMenu(self)
     # @param: self:object
     # @return: none
     # Assembles the menubar with all its compenents
     def makeMenu(self):
          self.menubar = Menu(self.master)
          goto = Menu(self.menubar)
          # goto.add_command(label = "Exit Program", command = self.exit())
          goto.add_command(label = "Combine Files", command = self.newCombine)
          goto.add_command(label = "Check File", command = self.newCompare)
          self.menubar.add_cascade(label = "Go to...", menu=goto)
          self.master.config(menu=self.menubar)
          user = Menu(self.menubar)
          user.add_command(label = "Info", command = self.seeUserInfo)
          user.add_command(label = "Log out", command = self.logout)
          self.menubar.add_cascade(label = "User", menu = user)
     
     # seeUserInfo(Self)
     # @param: self:object
     # @return: none
     # Displays currently login email in a message box
     def seeUserInfo(self):
          login = self.user
          message = "You're currently logged in as "+login
          box.showinfo("User Info", message)
          
     # logout(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and set new window to Login page
     def logout(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Login(self.newWindow)
     
     # newCombine(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and set new window to Combine page
     def newCombine(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Combine(self.newWindow, self.user, self.pw)
          
     # newCompare(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets new window to Compare page
     def newCompare(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Compare(self.newWindow, self.user, self.pw)
          
# Combine page
class Combine:
     
     #Constructor
     # __init__(self, master, user, pw)
     # @param: self:object, master:Tkinter, user:str, pw:str
     # @return: none
     # Constructs the Combine page with all its features
     def __init__(self, master, user, pw):
          self.master = master
          master.title("Combine Files")
          self.user = user
          self.pw = pw
          
          self.f1_btn = Button(master, text = "Choose File 1", command = self.openFile1)
          self.f1_btn.pack()
          
          self.f2_btn = Button(master, text = "Choose File 2", command = self.openFile2)
          self.f2_btn.pack()
          
          self.combine_btn = Button(master, text = "Combine", command = self.checkValid)
          self.combine_btn.pack()
          
          self.makeMenu()
          
     # makeMenu(self)
     # @param: self:object
     # @return: none
     # Assembles the menu with all is features
     def makeMenu(self):
          self.menubar = Menu(self.master)
          goto = Menu(self.menubar)
          goto.add_command(label = "Choose Function", command = self.toChoose)
          goto.add_command(label = "Combine Files", command = self.newCombine)
          goto.add_command(label = "Check File", command = self.newCompare)
          goto.add_command(label = "Previous Page", command = self.toChoose)
          self.menubar.add_cascade(label = "Go to...", menu=goto)
          self.master.config(menu=self.menubar)
          user = Menu(self.menubar)
          user.add_command(label = "Info", command = self.seeUserInfo)
          user.add_command(label = "Log out", command = self.logout)
          self.menubar.add_cascade(label = "User", menu = user)

     # checkCat(self, p)
     # @param: self:object, p:str
     # @return: valCat:boolean
     # Checks if the files the user chose have 
     def checkCat(self, p):
          book = load_workbook(p)
          sheet = book.active
          lastCol = sheet.max_column
          lastRow = sheet.max_row
          cat = []
          for i in range(1, lastCol+1):
               name = str(sheet.cell(row = 1, column = i).value)
               cat.append(name)
          valCat = False
          if cat[1].lower() != "first" and cat[1].lower() != "first name":    #If program does not have a first name category as the second category
               box.showerror("Missing Category","Please make sure your file has a first name category and is set as the second category.")
          elif cat[0].lower() != "last" and cat[0].lower() != "last name":    #If program does not have a last name catetory as the first category
               box.showerror("Missing Category","Please make sure your file has a last name category and is set as the first category.")
          else:
               valCat = True
          return valCat
          
     # openFile1(self)
     # @param: self:object
     # @return: none
     # Allows user to choose the first file, and checks if the file is in .xlsx format
     def openFile1(self):
          self.path1 = askopenfilename()
          length = len(self.path1)
          
          if self.path1[length-4:] != "xlsx":
               box.showerror("Invalid File Format", "Please choose a file with the .xlsx format.")
               self.val1 = False
          else:
               valCat1 = self.checkCat(self.path1)
               if valCat1 == True:
                    self.val1 = True
               else:
                    self.val1 = False

     # openFile2(self)
     # @param: self:object
     # @return: none
     # Allows user to choose the second file, and checks if the file is in .xlsx format
     def openFile2(self):
          self.path2 = askopenfilename()
          length = len(self.path2)
          if self.path2[length-4:] != "xlsx":
               box.showerror("Invalid File Format", "Please choose a file with the .xlsx format.")
               self.val2 = False
          else:
               valCat2 = self.checkCat(self.path2)
               if valCat2 == True:
                    self.val2 = True
               else:
                    self.val2 = False
          
     # checkValid(self)
     # @param: self:object
     # @return: none
     # Checks if both files are valid .xlsx files; if yes, continues to combine, if no, shows error message box
     def checkValid(self):
          self.valid = False
          if self.val1 == True and self.val2 == True:
               self.valid == True
               self.combineFiles()
          else:
               box.showerror("Invalid Files", "Please make sure you selected two files in .xlsx format and have the correct order of categories!")
                  
     # combineFiles(self)
     # @param: self:object
     # @return: none
     # Combines the two files together
     def combineFiles(self):
          book1 = load_workbook(self.path1)
          sheet1 = book1.active
          lastCol1 = sheet1.max_column
          lastRow1 = sheet1.max_row

          self.categories = []
          ln1 = []
          
          # Reads in category names
          for i in range(1, lastCol1+1):
              name = sheet1.cell(row = 1, column = i).value
              name = str(name)
              self.categories.append(name)
              
          # Reads in all data from Sheet 1
          for i in range(2, lastRow1+1):
               data = []
               for j in range(1, lastCol1+1):
                    info = sheet1.cell(row = i, column = j).value
                    info = str(info)
                    if info == 'None':
                         info = 'N/A'
                    data.append(info)
               ln1.append(data)

          # Loading Sheet 2
          book2 = load_workbook(self.path2)
          sheet2 = book2.active
          lastCol2 = sheet2.max_column
          lastRow2 = sheet2.max_row
          cat2 = []
          ln2 = []
          
          # Reads in all categories of sheet 2
          for i in range(1, lastCol2+1):
               name = sheet2.cell(row = 1, column = i).value
               name = str(name)
               cat2.append(name)
               
          # Reads in all datas of Sheet 2
          for i in range(2, lastRow2+1):
               data = []
               for j in range(1, lastCol2+1):
                    info = sheet2.cell(row = i, column = j).value
                    info = str(info)
                    if info == 'None':
                         info = "N/A"
                    data.append(info)
               ln2.append(data)
               
          # Checks for relapses in categories
          for i in range(0,len(cat2)):
               if cat2[i] not in self.categories:
                    self.categories.append(cat2[i])

          self.all = []
          # Adding all info from Sheet 1 and 2 to master list
          for i in range(0,len(ln1)):
               self.all.append(ln1[i])
          for i in range(0, len(ln2)):
               self.all.append(ln2[i])
          
          #Checks for repeated names and combines the information
          remove = []
          for i in range(0,len(ln2)):
               for j in range(0, len(ln1)):
                    if ln2[i][0] == ln1[j][0] and ln2[i][1] == ln1[j][1]:
                         tmp = []
                         for k in range(0,len(ln1[j])):
                              tmp.append(ln1[j][k])
                         for k in range(2, len(ln2[i])):
                              tmp.append(ln2[i][k])
                         self.all.remove(ln1[j])
                         self.all.remove(ln2[i])
                         self.all.append(tmp)
          
          # Put data in correct place
          for i in range(0, len(self.all)):
               # Ln1 data should not have the issue of misplacing info, only Ln2 data
               if len(self.all[i]) < len(self.categories) and self.all[i] in ln2:
                    ind = ln2.index(self.all[i])
                    # for j in range(0,len(ln2[ind])):
                    if cat2[j]!=self.categories[j]:
                         ind = self.categories.index(cat2[j])
                         for k in range(j, ind):
                              self.all[i].insert(k, 'N/A')
               elif len(self.all[i]) < len(self.categories) and self.all[i] in ln1:
                    goalLen = len(self.categories)
                    current = len(self.all[i])
                    # diff = goalLen-current
                    for k in range(current, goalLen):
                         self.all[i].append("N/A")
                         
          # Sort data according to last names                 
          self.all.sort()
          
          # Writes combined file 
          workbook = xlsxwriter.Workbook('Combined_File.xlsx')
          worksheet = workbook.add_worksheet()
          for i in range(0,len(self.categories)):
               worksheet.write(0, i, self.categories[i])
          for i in range(1, len(self.all)+1):
               for j in range(0,len(self.all[i-1])):
                    worksheet.write(i,j, self.all[i-1][j])
          workbook.close()
          
          # Message for user
          box.showinfo("Message", "File saved in program's folder.")
          
          self.toChoose()
          
     # seeUserInfo(self)
     # @param: self:object
     # @return: none
     # Displays currently login email in a message box
     def seeUserInfo(self):
          login = self.user
          message = "You're currently logged in as "+login
          box.showinfo("User Info", message)
          
     # logout(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and set new window to Login page
     def logout(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Login(self.newWindow)
          
     # toChoose(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and set new window to Choose page
     def toChoose(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Choose(self.newWindow, self.user, self.pw)
          
     # newCombine(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets new window to Combine page
     def newCombine(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Combine(self.newWindow)
          
     # newCompare(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets new window to Compare page
     def newCompare(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Compare(self.newWindow)
          
           
# Compare page
class Compare:
     
     # Constructor
     # __init__(self, master, user, pw)
     # @param: self:object, master:Tkinter, user:str, pw:str
     # @return: none
     # Constructs and compiles all features of Compare page
     def __init__(self,master, user, pw):
          self.master = master
          master.title("Find Missing Info")
          self.user = user
          self.pw = pw
          
          self.msg = Label(master, text = "Please choose a file to check for missing information.")
          self.msg.pack()
          
          self.f_btn = Button(master, text = "Choose file", command = self.getPath)
          self.f_btn.pack()
          
          self.missing = []
          
          self.check_btn = Button(master, text = "Check", command = self.checkValid)
          self.check_btn.pack()
          
          self.makeMenu()
          
     # makeMenu(self)
     # @param: self:object
     # @return: none
     # Assembles menubar with all its features
     def makeMenu(self):
          self.menubar = Menu(self.master)
          goto = Menu(self.menubar)
          goto.add_command(label = "Choose Function", command = self.toChoose)
          goto.add_command(label = "Combine Files", command = self.newCombine)
          goto.add_command(label = "Check File", command = self.newCompare)
          goto.add_command(label = "Previous Page", command = self.toChoose)
          self.menubar.add_cascade(label = "Go to...", menu=goto)
          self.master.config(menu=self.menubar)
          user = Menu(self.menubar)
          user.add_command(label = "Info", command = self.seeUserInfo)
          user.add_command(label = "Log out", command = self.logout)
          self.menubar.add_cascade(label = "User", menu = user)
     
     # getPath(self)
     # @param: self:object
     # @return: none
     # Gets file path and checks if its format is .xlsx
     def getPath(self):
          self.path = askopenfilename()
          self.valid = False
          length = len(self.path)
          if self.path[length-4:] != "xlsx":
               box.showerror("Invalid File Format", "Please choose a file with the .xlsx format.")
          else:
               self.valid = True
     
     # checkValid(self)
     # @param: self:object
     # @return: none
     # Checks if the file has the correct format, and if it has a last name, then first name category
     def checkValid(self):
          if self.valid == False:  #If the format is not .xlsx
               box.showerror("Invalid File Format", "Please choose a file with the .xlsx format!")
          else:     #If file format is correct
               book = load_workbook(self.path)
               self.sheet = book.active
               self.lastCol = self.sheet.max_column
               self.lastRow = self.sheet.max_row     
     
               cat = []
               for i in range(1,self.lastCol+1):  #Reading in data from selected xlsx file
                    tmp = self.sheet.cell(row = 1, column = i).value
                    tmp = str(tmp)
                    cat.append(tmp)
               
               if cat[1].lower() != "first" and cat[1].lower() != "first name":    #If program does not have a first name category as the second category
                    box.showerror("Missing Category","Please make sure your file has a first name category and is set as the second category.")
                    self.getPath()
                    self.checkValid()
               elif cat[0].lower() != "last" and cat[0].lower() != "last name":    #If program does not have a last name catetory as the first category
                    box.showerror("Missing Category","Please make sure your file has a last name category and is set as the first category.")
                    self.getPath()
                    self.checkValid()
               else:
                    self.checkMissing()
               
     # checkMissing(self)
     # @param: self:object
     # @return: none
     # Checks for missing information (missing cells or cells with 'N/A')
     def checkMissing(self):
          for i in range(2, self.lastRow+1):
               emp = 0
               data = []
               for j in range(1, self.lastCol+1):
                    tmp = self.sheet.cell(row = i, column = j).value
                    tmp = str(tmp)
                    add = True
                    if tmp == 'N/A' or tmp == 'None':  #If cell does not have information
                         if emp == 0:   #If information in first name or last name is missing
                              last = str(self.sheet.cell(row=i,column=1).value)
                              first = str(self.sheet.cell(row = i, column = 2).value)
                              if last == 'None' or first == 'None' or last =='N/A' or first == 'N/A':
                                   add = False
                                   break
                              else:     #If first name and last name information is present
                                   name = []
                                   name.append(first)
                                   name.append(last)
                                   data.append(name)
                                   emp += 1
                         cat = str(self.sheet.cell(row = 1, column = j).value)
                         data.append(cat)
               if add == True:
                    self.missing.append(data)
          
          self.newInfoMissing()
          
     # newInfoMissing(self)
     # @param: self:object
     # @return: none
     # If missing information is present, withdraws current window and set to InfoMissing page;
     # If there's no missing information present in list, outputs a message box and informs the user about it, and returns to Choose page
     def newInfoMissing(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          for i in range(0,len(self.missing)):
               if len(self.missing[i]) != 0:
                    self.root = InfoMissing(self.newWindow, self.missing, self.user, self.pw)
                    break
               elif len(self.missing[i]) == 0 and i == len(self.missing)-1:
                    box.showinfo("All caught up!", "There is no missing information in your file.")
                    self.root = Choose(self.newWindow, self.user, self.pw)

     # seeUserInfo(self)
     # @param: self:object
     # @return: none
     # Displays current login email in a message box
     def seeUserInfo(self):
          login = self.user
          message = "You're currently logged in as "+login
          box.showinfo("User Info", message)
          
     # logout(self)
     # @param: self:object
     # @return: none
     # Withdraws current page and set back to Login page
     def logout(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Login(self.newWindow)
          
     # toChooose(self)
     # @param: none
     # @return: none
     # Withdraws current page and set to Choose page
     def toChoose(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Choose(self.newWindow, self.user, self.pw)
          
     # newCombine(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets to Combine page
     def newCombine(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Combine(self.newWindow, self.user, self.pw)
          
     # newCompare(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets to Compare page
     def newCompare(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Compare(self.newWindow, self.user, self.pw)
          
# InfoMissing page
class InfoMissing:
     # Constructor
     # __init__(self, master, missing, user, pw)
     # @param: self:object, master:Tkinter, missing:list, user:str, pw:str
     # @return: none
     # Constructs InfoMissing page with all its features
     def __init__(self, master, missing, user, pw):
          self.master = master
          master.title("Missing Information")
          self.user = user
          self.pw = pw
          
          self.missing = missing
           
          self.name_lbl = Label(master, text = "Name")
          self.name_lbl.grid(row = 1, column = 0)
          
          self.infoNeeded_lbl = Label(master, text = "Info Needed")
          self.infoNeeded_lbl.grid(row = 1, column = 1)
          
          self.checkBox()
          
          self.email_btn = Button(master, text = "Proceed to Email", command = self.next)
          self.email_btn.grid(row = self.dataRow+2, column = 3)
          
          self.makeMenu()

     # makeMenu(self)
     # @param: self:object
     # @return: none
     # Assembles the menu with all its features
     def makeMenu(self):
          self.menubar = Menu(self.master)
          goto = Menu(self.menubar)
          goto.add_command(label= "Choose Function", command = self.toChoose)
          goto.add_command(label = "Combine Files", command = self.newCombine)
          goto.add_command(label = "Check File", command = self.newCompare)
          goto.add_command(label = "Previous Page", command = self.newCompare)
          self.menubar.add_cascade(label = "Go to...", menu=goto)
          self.master.config(menu=self.menubar)
          user = Menu(self.menubar)
          user.add_command(label = "Info", command = self.seeUserInfo)
          user.add_command(label = "Log out", command = self.logout)
          self.menubar.add_cascade(label = "User", menu = user)
          
     # checkBox(self)
     # @param: self:object
     # @return: none
     # Puts names in the missing list into a list box; puts missing assignments the students have beside the list box options
     def checkBox(self):
          self.names = []
          info = []
          for i in range(0, len(self.missing)):
               self.names.append(self.missing[i][0])
               tmp = []
               for j in range(1,len(self.missing[i])):
                    tmp.append(self.missing[i][j])
               info.append(tmp)
          
          for i in range(0, len(self.names)):
               self.names[i] = self.names[i][0]+" "+self.names[i][1]
          
          for i in range(0,len(self.names)):
               self.listbox = Listbox(self.master, selectmode = MULTIPLE)
               self.listbox.grid(row = 2, column = 0, sticky = 'W', rowspan = len(self.names))
               for j in range(0, len(self.names)):
                    self.listbox.insert(i, self.names[j])
               data = info[i]
               tmp = ""
               for j in range(0,len(data)):
                    tmp = tmp+data[j]
                    if j != len(data)-1:
                         tmp = tmp + ", "
               d = Label(self.master, text = tmp)
               d.grid(row = i+2, rowspan = 1, column = 1, sticky = 'W')
               
          self.dataRow = len(self.names)
          
     # select(self)
     # @param: self:object
     # @return: none
     # Looks at selected options in the list box and puts them in a list
     def select(self):
          self.emailList = []
          selections = self.listbox.curselection()
          selections = [int(x) for x in selections]
          values = [self.emailList.append(self.names[x]) for x in selections]
     
     # next(self)
     # @param: self:object
     # @return: none
     # If the selection is empty, program asks if the user is sure about their choice
     # If the selection is not empty, proceeds to Email page
     def next(self):
          self.select()
          if len(self.emailList) == 0:
               check = box.askyesno("Empty Selections", "You did not select any recipients. Are you sure about your choice?")
               if check == True:
                    self.toChoose()
          else:
               self.master.withdraw()
               self.newWindow = Toplevel(self.master)
               self.root = Email(self.newWindow, self.emailList, self.user, self.pw, self.missing)

     # seeUserInfo(self)
     # @param: self:object
     # @return: none
     # Displays currently login email in a message box
     def seeUserInfo(self):
          login = self.user
          message = "You're currently logged in as "+login
          box.showinfo("User Info", message)
          
     # logout(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets back go Login page
     def logout(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Login(self.newWindow)
          
     # toChoose(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets back to Choose page
     def toChoose(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Choose(self.newWindow, self.user, self.pw)
          
     # newCombine(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets page to Combine page
     def newCombine(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Combine(self.newWindow, self.user, self.pw)
          
     # newCompare(self)
     # @param: self:object
     # @return: none
     # Withdraws current window and sets page to Compare page
     def newCompare(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Compare(self.newWindow, self.user, self.pw)
          
# Email page
class Email:
     
     # Constructor
     # __init__(self, master, emailList, user, pw, missing)
     # @param: self:object, master:Tkinter, emailList:list, user:str, pw:str, missing:2Dlist
     # @return: none
     # Constructs Email page with all its features in place
     def __init__(self, master, emailList, user, pw, missing):
          self.master = master
          master.title("Email")
          self.user = user
          self.pw = pw
          self.toList = emailList
          self.missing = missing
          
          self.recipient_lbl = Label(master, text = "Recipients:")
          self.recipient_lbl.grid(row = 0, column = 0)
 
          for i in range(0, len(self.toList)):
               self.recipient_list=Label(master, text=self.toList[i])
               self.recipient_list.grid(row=i, column=1)
          
          self.subject_lbl = Label(master, text='Subject:')
          self.subject_lbl.grid(row=i+1, column=0)
          
          self.subject_ent = Entry(master)
          self.subject_ent.grid(row=i+1,column=1)
          
          self.message_lbl=Label(master, text="Message:")
          self.message_lbl.grid(row =i+3, column = 0)
          
          self.message_ent = Entry(master)
          self.message_ent.grid(row = i+3,rowspan = 1, column = 1)
          
          self.send_btn = Button(master, text = "Send", command = self.getAddress)
          self.send_btn.grid(row = i+4, column = 1)
     
          self.makeMenu()
     
     # makeMenu(self)
     # @param: self:object
     # @return: none
     # Assembles the menubar with all its features
     def makeMenu(self):
          self.menubar = Menu(self.master)
          goto = Menu(self.menubar)
          # goto.add_command(label = "Exit Program", command = self.exit())
          goto.add_command(label= "Choose Function", command = self.toChoose)
          goto.add_command(label = "Combine Files", command = self.newCombine)
          goto.add_command(label = "Check File", command = self.newCompare)
          goto.add_command(label = "Previous Page", command = self.newInfoMissing)
          self.menubar.add_cascade(label = "Go to...", menu=goto)
          self.master.config(menu=self.menubar)
          user = Menu(self.menubar)
          user.add_command(label = "Info", command = self.seeUserInfo)
          user.add_command(label = "Log out", command = self.logout)
          self.menubar.add_cascade(label = "User", menu = user)
     
     # getAddress(self)
     # @param: self:object
     # @return: none
     # Extracts the corresponding email addresses from the Master Directory 
     def getAddress(self):
          f = open("Master_Directory.xlsx")
          book = load_workbook("Master_Directory.xlsx")
          sheet = book.active
          lastRow = sheet.max_row
          last = []
          first = []
          email = []
          for i in range(1, lastRow+1):      # Puts all the information in its corresponding lists
               lastName = str(sheet.cell(row = i, column = 1).value)
               firstName = str(sheet.cell(row = i, column = 2).value)
               address = str(sheet.cell(row = i, column = 3).value)
               last.append(lastName)
               first.append(firstName)
               email.append(address)
               
          for i in range(0, len(self.toList)):
               self.toList[i].split()

          self.addresses = []
          found = False
          foundNum = 0
          pointer = 0
          for i in range(0,len(self.toList)):     # Going through the list of names need to be found
               ref = self.toList[i]     # Sets reference point to the name looking for
               while found == False and pointer < len(email):    # While name not found and have not went through the whole list
                    if str(first[pointer]+' '+last[pointer]) == ref:  # If the name is found
                         self.addresses.append(email[pointer])
                         found == True
                         foundNum += 1
                         break
                    elif pointer == len(email)-1 and found == False:
                         msg = "The email address for {} is not found in Master_Directory.xlsx. Please make sure the information is present in the file.".format(ref)
                         box.showerror("Info not found", msg)
                         found = True
                    else:
                         found == False
                         pointer += 1
          if foundNum == len(self.toList):
               self.sendEmail()
          
     # sendEmail(self)
     # @param: self:object
     # @return: none
     # Sends the email to the recipient(s)
     def sendEmail(self):
          mail = smtplib.SMTP('smtp.gmail.com', 587)
          mail.ehlo()
          mail.starttls()

          mail.login(self.user, self.pw)
          recipients = self.addresses
          subject = self.subject_ent.get()
          message_text = self.message_ent.get()
          message = "From: %s\r\n" % self.user + "To: %s\r\n" % recipients + "Subject: %s\r\n" % subject + "\r\n" + message_text
          mail.sendmail(self.user, recipients, message)
          mail.close()
          
          box.showinfo("Message", "Email sent!")
          
          self.clear()
     
     # clear(self)
     # @param: self:object
     # @return: none
     # Clears the entry boxes after email is sent
     def clear(self):
          self.subject_ent.delete(0,"end")
          self.message_ent.delete(0,"end")
     
     # seeUserInfo(self)
     # @param: self:object
     # @return: none
     # Displays currently login email in a message box
     def seeUserInfo(self):
          login = self.user
          message = "You're currently logged in as "+login
          box.showinfo("User Info", message)
          
     # logout(self)
     # @param: self:object
     # @return: none
     # Withdraws current page and sets back to login page
     def logout(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Login(self.newWindow)
          
     # toChoose(self)
     # @param: self:object
     # @return: none
     # Withdraws current page and sets back to Choose page
     def toChoose(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Choose(self.newWindow, self.user, self.pw)
          
     # newCombine(self)
     # @param: self:object
     # @return: none
     # Withdraws current page and sets back to Combine page
     def newCombine(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Combine(self.newWindow, self.user, self.pw)
          
     # newCompare(self)
     # @param: self:object
     # @return:none
     # Withdraws current page and sets back to Compare page
     def newCompare(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = Compare(self.newWindow, self.user, self.pw)
          
     # newInfoMissing(self)
     # @param: self:object
     # @return: none
     # Withdraws current page and sets back to InfoMissing page
     def newInfoMissing(self):
          self.master.withdraw()
          self.newWindow = Toplevel(self.master)
          self.root = InfoMissing(self.newWindow, self.missing, self.user, self.pw)

          
window = Tk()
# Starting with Login page
gui = Login(window)
# Run program
window.mainloop()